from openpyxl import Workbook
from openpyxl import load_workbook

workb = Workbook()
try:
    workb = load_workbook("binance_exportx.xlsx") #original excel file
except:
    workb = Workbook() # if file doesn't exists, create new
ws = workb.active # to get the active cell

def readEntireData():
    """
    Reads the entire excel file. 
    returns a list.
    
    """
    data = tuple(ws.rows)
    for each in data:
        temp = []
        for values in each:
            temp.append(values.value)
        yield temp

def read_formatted_data(rawData):
    """
    str(rawData) -> list
    Converts raw excel data into formatted reusable list of dicionaty data
    returns a list
    Example: [{title:"some title",type:"c",price:456,date:"17/01/2021"}]
    """
    result = []
    for each in rawData:
        temp = {}
        types = ["title","type","price","date"]
        for i in range(len(each)):
            temp.update({types[i]:each[i]})
        result.append(temp)
    return result

def insertData(price,typee,title):
    """
    str(price) -> int
    str(typee) -> str
    str(title) -> str
    Adds the transaction into the excel database.
    Returns nothing.
    """
    length = len(list(ws.rows))
    ws.cell(row=length+1,column=1,value=title)
    ws.cell(row=length+1,column=2,value=typee)
    ws.cell(row=length+1,column=3,value=price)
    ws.cell(row=length+1,column=4,value=return_today())
    workb.save("database.xlsx")

if __name__ == "__main__":
    print(list(readEntireData()))
    workb.save("database.xlsx")